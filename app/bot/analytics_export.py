from datetime import datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from app.bot.analytics import sales_rows
from app.db.session import SessionLocal
from app.models import (
    Club, Discrepancy, Employee, Guest, GuestTelegram, InventoryBalance,
    InventoryOperation, Product, SalaryPeriod, Shift, Writeoff, WriteoffItem,
    WriteoffStatus,
)
from sqlalchemy import func, select


def _dec(v):
    return Decimal(str(v or 0))


def _style(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        values = [str(c.value) if c.value is not None else "" for c in col]
        width = min(max(max(map(len, values), default=10) + 2, 10), 45)
        ws.column_dimensions[get_column_letter(col[0].column)].width = width


def _sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    _style(ws)
    return ws


async def build_excel(days: int = 30, start: datetime | None = None, end: datetime | None = None) -> Path:
    """Build Excel export for a period.

    Backwards compatible with the analytics screen (days=30) and supports
    exact UTC windows used by the ежедневный отчёт владельцу.
    """
    if start is None or end is None:
        end = datetime.now(timezone.utc)
        start = end - timedelta(days=days)
    else:
        if start.tzinfo is None:
            start = start.replace(tzinfo=timezone.utc)
        if end.tzinfo is None:
            end = end.replace(tzinfo=timezone.utc)
    sales = await sales_rows(start, end)

    async with SessionLocal() as session:
        shift_rows = (await session.execute(
            select(Shift, Employee, Club)
            .outerjoin(Employee, Employee.id == Shift.employee_id)
            .join(Club, Club.id == Shift.club_id)
            .where(Shift.started_at >= start, Shift.started_at <= end)
            .order_by(Shift.started_at)
        )).all()
        balance_rows = (await session.execute(
            select(Club.name, Product.name, InventoryBalance.quantity, InventoryBalance.min_stock)
            .join(Club, Club.id == InventoryBalance.club_id)
            .join(Product, Product.id == InventoryBalance.product_id)
            .order_by(Club.name, Product.name)
        )).all()
        op_rows = (await session.execute(
            select(InventoryOperation.created_at, Club.name, Product.name,
                   InventoryOperation.operation_type, InventoryOperation.quantity,
                   Employee.full_name, InventoryOperation.source, InventoryOperation.source_id,
                   InventoryOperation.comment)
            .join(Club, Club.id == InventoryOperation.club_id)
            .join(Product, Product.id == InventoryOperation.product_id)
            .outerjoin(Employee, Employee.id == InventoryOperation.employee_id)
            .where(InventoryOperation.created_at >= start, InventoryOperation.created_at <= end)
            .order_by(InventoryOperation.created_at)
        )).all()
        writeoffs = (await session.execute(
            select(Writeoff.created_at, Club.name, Employee.full_name, Writeoff.status,
                   WriteoffItem.quantity, Product.name, Writeoff.comment)
            .join(Club, Club.id == Writeoff.club_id)
            .join(Employee, Employee.id == Writeoff.employee_id)
            .join(WriteoffItem, WriteoffItem.writeoff_id == Writeoff.id)
            .join(Product, Product.id == WriteoffItem.product_id)
            .where(Writeoff.created_at >= start, Writeoff.created_at <= end)
            .order_by(Writeoff.created_at)
        )).all()
        discrepancies = (await session.execute(
            select(Discrepancy.created_at, Club.name, Product.name, Employee.full_name,
                   Discrepancy.quantity_difference, Discrepancy.amount_difference,
                   Discrepancy.status, Discrepancy.reason, Discrepancy.resolution_comment)
            .join(Club, Club.id == Discrepancy.club_id)
            .join(Product, Product.id == Discrepancy.product_id)
            .outerjoin(Employee, Employee.id == Discrepancy.employee_id)
            .where(Discrepancy.created_at >= start, Discrepancy.created_at <= end)
            .order_by(Discrepancy.created_at)
        )).all()
        salary = (await session.execute(
            select(SalaryPeriod, Employee.full_name)
            .join(Employee, Employee.id == SalaryPeriod.employee_id)
            .where(SalaryPeriod.date_from <= end.date(), SalaryPeriod.date_to >= start.date())
            .order_by(Employee.full_name, SalaryPeriod.date_from)
        )).all()
        clients = (await session.execute(
            select(Guest.langame_guest_id, Guest.fio, Guest.phone, Guest.is_temp,
                   Guest.is_virtual, GuestTelegram.telegram_user_id,
                   GuestTelegram.marketing_consent, GuestTelegram.marketing_consent_at)
            .outerjoin(GuestTelegram, GuestTelegram.guest_id == Guest.id)
            .order_by(Guest.fio)
        )).all()

    wb = Workbook()
    wb.remove(wb.active)

    total_sales = Decimal("0")
    total_units = Decimal("0")
    by_shift = {}
    for row in sales:
        if int(row.get("cancel", 0) or 0) == 1 or row.get("working_shift_id") is None:
            continue
        try:
            sid = int(row["working_shift_id"])
            qty = _dec(row.get("count"))
            amount = _dec(row.get("price_sale")) * qty
        except Exception:
            continue
        total_sales += amount
        total_units += qty
        x = by_shift.setdefault(sid, [Decimal("0"), Decimal("0")])
        x[0] += amount
        x[1] += qty

    _sheet(wb, "Сводка", ["Показатель", "Значение"], [
        ("Период с", start.strftime("%Y-%m-%d %H:%M")),
        ("Период по", end.strftime("%Y-%m-%d %H:%M")),
        ("Товарные продажи LANGAME, ₽", float(total_sales)),
        ("Продано единиц", float(total_units)),
        ("Смен", len(shift_rows)),
        ("Закрытых смен", sum(1 for sh, _, _ in shift_rows if sh.status == "closed")),
        ("Низких остатков", sum(1 for _, _, q, m in balance_rows if _dec(m) > 0 and _dec(q) <= _dec(m))),
        ("Расхождений", len(discrepancies)),
        ("Одобренных списаний, строк", sum(1 for x in writeoffs if x[3] == WriteoffStatus.APPROVED.value)),
        ("Клиентов", len(clients)),
        ("Telegram-привязок", sum(1 for x in clients if x[5] is not None)),
        ("Marketing consent", sum(1 for x in clients if x[6] is True)),
    ])

    _sheet(wb, "Продажи", ["Дата", "Смена LANGAME", "Количество", "Цена", "Сумма", "Отменено"], [
        (row.get("date") or row.get("created_at") or "", row.get("working_shift_id"),
         float(_dec(row.get("count"))), float(_dec(row.get("price_sale"))),
         float(_dec(row.get("count")) * _dec(row.get("price_sale"))), int(row.get("cancel", 0) or 0))
        for row in sales
    ])

    admin_rows = []
    for sh, emp, club in shift_rows:
        sale, units = by_shift.get(int(sh.langame_shift_id), (Decimal("0"), Decimal("0")))
        hours = Decimal("0")
        if sh.ended_at and sh.ended_at > sh.started_at:
            hours = Decimal(str((sh.ended_at - sh.started_at).total_seconds())) / Decimal("3600")
        admin_rows.append((emp.full_name if emp else "Не привязан", club.name, sh.langame_shift_id,
                           sh.started_at, sh.ended_at, sh.status, float(hours), float(sale), float(units),
                           float(_dec(sh.cash_difference))))
    _sheet(wb, "Смены", ["Администратор", "Клуб", "Смена LANGAME", "Начало", "Конец", "Статус", "Часы", "Продажи ₽", "Единиц", "Разница кассы ₽"], admin_rows)

    admin_summary = {}
    for r in admin_rows:
        name = r[0]
        x = admin_summary.setdefault(name, [0, Decimal("0"), Decimal("0"), Decimal("0"), Decimal("0")])
        x[0] += 1; x[1] += _dec(r[6]); x[2] += _dec(r[7]); x[3] += _dec(r[8]); x[4] += _dec(r[9])
    _sheet(wb, "Администраторы", ["Администратор", "Смен", "Часы", "Продажи ₽", "Единиц", "Продажи/час ₽", "Разница кассы ₽"], [
        (name, x[0], float(x[1]), float(x[2]), float(x[3]), float(x[2] / x[1]) if x[1] else 0, float(x[4]))
        for name, x in sorted(admin_summary.items(), key=lambda kv: kv[1][2], reverse=True)
    ])

    _sheet(wb, "Остатки бара и снеков", ["Клуб", "Товар", "Остаток", "Минимум", "Статус"], [
        (club, product, float(_dec(qty)), float(_dec(minimum)), "НИЗКИЙ" if _dec(minimum) > 0 and _dec(qty) <= _dec(minimum) else "OK")
        for club, product, qty, minimum in balance_rows
    ])
    _sheet(wb, "Операции бар и снеки", ["Дата", "Клуб", "Товар", "Тип", "Количество", "Администратор", "Источник", "ID источника", "Комментарий"], op_rows)
    _sheet(wb, "Списания", ["Дата", "Клуб", "Администратор", "Статус", "Количество", "Товар", "Комментарий"], writeoffs)
    _sheet(wb, "Расхождения", ["Дата", "Клуб", "Товар", "Администратор", "Разница", "Сумма ₽", "Статус", "Причина", "Комментарий решения"], discrepancies)
    _sheet(wb, "Зарплата", ["Администратор", "С", "По", "База ₽", "Бонус ₽", "Итого ₽", "Статус"], [
        (name, p.date_from, p.date_to, float(_dec(p.base_amount)), float(_dec(p.bonus_amount)), float(_dec(p.total_amount)), p.status)
        for p, name in salary
    ])
    _sheet(wb, "Клиенты", ["LANGAME ID", "ФИО", "Телефон", "Временный", "Виртуальный", "Telegram ID", "Marketing consent", "Consent at"], clients)

    out_dir = Path("/tmp/langame_exports")
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"langame_analytics_{datetime.now(timezone.utc):%Y%m%d_%H%M%S}.xlsx"
    wb.save(path)
    return path
