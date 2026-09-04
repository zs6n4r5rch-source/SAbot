from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO

from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select

from app.db.session import SessionLocal
from app.models import (
    AuditLog, Club, Discrepancy, Employee, InventoryBalance, NonMonetaryBonus,
    SalaryPeriod, SalaryViolation, Shift, TelegramUser, Writeoff, WriteoffItem,
)
from app.services.langame import langame_client, LangameAPIError
from app.bot.analytics import sales_rows

router = APIRouter(prefix="/api/statistics", tags=["statistics"])


def dec(v):
    return float(v or 0)


def period_bounds(days: int):
    end = datetime.now(timezone.utc)
    if days <= 0:
        return datetime(2000, 1, 1, tzinfo=timezone.utc), end, "за всё время"
    return end.replace(hour=0, minute=0, second=0, microsecond=0), end, f"последние {days} дней"


async def collect(days: int):
    start, end, label = period_bounds(days)
    sales = await sales_rows(start, end)
    valid_sales = [r for r in sales if int(r.get("cancel", 0) or 0) != 1]
    sales_total = Decimal("0")
    units = Decimal("0")
    for r in valid_sales:
        try:
            units += Decimal(str(r.get("count", 0) or 0))
            sales_total += Decimal(str(r.get("price_sale", 0) or 0)) * Decimal(str(r.get("count", 0) or 0))
        except Exception:
            continue

    async with SessionLocal() as session:
        shift_rows = (await session.execute(
            select(Shift, Employee, Club)
            .outerjoin(Employee, Employee.id == Shift.employee_id)
            .outerjoin(Club, Club.id == Shift.club_id)
            .where(Shift.started_at >= start, Shift.started_at <= end)
            .order_by(Shift.started_at)
        )).all()
        violations = (await session.execute(
            select(SalaryViolation, Employee).join(Employee, Employee.id == SalaryViolation.employee_id)
            .where(SalaryViolation.created_at >= start, SalaryViolation.created_at <= end)
            .order_by(SalaryViolation.created_at)
        )).all()
        discrepancies = (await session.execute(
            select(Discrepancy, Employee).join(Employee, Employee.id == Discrepancy.employee_id)
            .where(Discrepancy.created_at >= start, Discrepancy.created_at <= end)
        )).all()
        writeoffs = (await session.execute(
            select(WriteoffItem, Writeoff, Employee)
            .join(Writeoff, Writeoff.id == WriteoffItem.writeoff_id)
            .join(Employee, Employee.id == Writeoff.employee_id)
            .where(Writeoff.created_at >= start, Writeoff.created_at <= end)
        )).all()
        salaries = (await session.execute(
            select(SalaryPeriod, Employee).join(Employee, Employee.id == SalaryPeriod.employee_id)
            .where(SalaryPeriod.date_from <= end.date(), SalaryPeriod.date_to >= start.date())
        )).all()
        bonuses = (await session.execute(
            select(NonMonetaryBonus, Employee).join(Employee, Employee.id == NonMonetaryBonus.employee_id)
            .where(NonMonetaryBonus.created_at >= start, NonMonetaryBonus.created_at <= end)
        )).all()
        inventory_critical = await session.scalar(select(func.count(InventoryBalance.id)).where(
            InventoryBalance.min_stock > 0, InventoryBalance.quantity <= InventoryBalance.min_stock
        )) or 0
        employees = (await session.execute(select(Employee).where(Employee.active.is_(True)).order_by(Employee.full_name))).scalars().all()

    sales_by_shift = {}
    for r in valid_sales:
        sid = r.get("working_shift_id")
        if sid is None:
            continue
        try:
            amount = Decimal(str(r.get("price_sale", 0) or 0)) * Decimal(str(r.get("count", 0) or 0))
            sales_by_shift[int(sid)] = sales_by_shift.get(int(sid), Decimal("0")) + amount
        except Exception:
            pass

    admins = {e.id: {"id": e.id, "name": e.full_name or f"Администратор #{e.id}", "shifts": 0, "hours": Decimal("0"), "sales": Decimal("0"), "cash": Decimal("0"), "violations": Decimal("0"), "salary": Decimal("0")} for e in employees}
    for sh, emp, club in shift_rows:
        if not emp or emp.id not in admins:
            continue
        st = admins[emp.id]
        st["shifts"] += 1
        if sh.ended_at and sh.ended_at > sh.started_at:
            st["hours"] += Decimal(str((sh.ended_at - sh.started_at).total_seconds())) / Decimal("3600")
        st["sales"] += sales_by_shift.get(int(sh.langame_shift_id), Decimal("0"))
        st["cash"] += Decimal(str(sh.cash_difference or 0))
    for v, emp in violations:
        if emp.id in admins:
            admins[emp.id]["violations"] += Decimal(str(v.amount or 0))
    for p, emp in salaries:
        if emp.id in admins:
            admins[emp.id]["salary"] += Decimal(str(p.total_amount or 0))

    admin_rows = []
    for st in admins.values():
        sph = st["sales"] / st["hours"] if st["hours"] else Decimal("0")
        admin_rows.append({**st, "sales_per_hour": sph})
    admin_rows.sort(key=lambda x: (x["sales"], x["hours"]), reverse=True)

    return {
        "period": label, "from": start.isoformat(), "to": end.isoformat(),
        "summary": {
            "sales": dec(sales_total), "units": dec(units), "sale_rows": len(valid_sales),
            "shifts": len(shift_rows), "closed_shifts": sum(1 for s, _, _ in shift_rows if s.status == "closed"),
            "admin_hours": dec(sum((x["hours"] for x in admin_rows), Decimal("0"))),
            "cash_difference": dec(sum((x["cash"] for x in admin_rows), Decimal("0"))),
            "penalties": dec(sum((x["violations"] for x in admin_rows), Decimal("0"))),
            "salary": dec(sum((x["salary"] for x in admin_rows), Decimal("0"))),
            "discrepancies": len(discrepancies),
            "discrepancy_amount": dec(sum((Decimal(str(d.amount_difference or 0)) for d, _ in discrepancies), Decimal("0"))),
            "writeoff_units": dec(sum((Decimal(str(wi.quantity or 0)) for wi, _, _ in writeoffs), Decimal("0"))),
            "critical_stock": int(inventory_critical),
            "violations_count": len(violations), "bonuses_count": len(bonuses),
        },
        "admins": [{k: (dec(v) if isinstance(v, Decimal) else v) for k, v in x.items()} for x in admin_rows],
        "sales": [{"date": r.get("date") or r.get("created_at") or r.get("sale_date"), "product": r.get("product_name") or r.get("name"), "quantity": dec(r.get("count", 0)), "price": dec(r.get("price_sale", 0)), "working_shift_id": r.get("working_shift_id")} for r in valid_sales],
        "violations": [{"date": v.created_at.isoformat() if v.created_at else None, "employee": e.full_name, "rule": v.rule_code, "amount": dec(v.amount), "premium_reduction_percent": dec(v.premium_reduction_percent), "dismissal_required": v.dismissal_required} for v, e in violations],
        "discrepancies_detail": [{"date": d.created_at.isoformat() if d.created_at else None, "employee": e.full_name, "amount": dec(d.amount_difference)} for d, e in discrepancies],
        "writeoffs": [{"date": w.created_at.isoformat() if w.created_at else None, "employee": e.full_name, "quantity": dec(wi.quantity), "status": w.status} for wi, w, e in writeoffs],
        "salaries": [{"from": p.date_from.isoformat(), "to": p.date_to.isoformat(), "employee": e.full_name, "base": dec(p.base_amount), "bonus": dec(p.bonus_amount), "total": dec(p.total_amount), "status": p.status} for p, e in salaries],
    }


@router.get("")
async def statistics(request: Request, days: int = 30):
    from app.webapp.app import current_user, owner_required
    user, _ = await current_user(request)
    owner_required(user)
    days = max(0, min(days, 36500))
    return await collect(days)


@router.get("/export")
async def export_statistics(request: Request, days: int = 30):
    from app.webapp.app import current_user, owner_required
    user, _ = await current_user(request)
    owner_required(user)
    days = max(0, min(days, 36500))
    data = await collect(days)
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"
    for row in [("Период", data["period"]), *[(k, v) for k, v in data["summary"].items()]]:
        ws.append(list(row))

    def sheet(name, rows):
        sh = wb.create_sheet(name)
        if not rows:
            sh.append(["Нет данных"])
            return
        headers = list(rows[0].keys())
        sh.append(headers)
        for row in rows:
            sh.append([row.get(h) for h in headers])
        for col in range(1, len(headers) + 1):
            sh.column_dimensions[get_column_letter(col)].width = min(max(len(str(sh.cell(1, col).value or "")) + 2, 12), 32)
        sh.freeze_panes = "A2"
    sheet("Администраторы", data["admins"])
    sheet("Продажи", data["sales"])
    sheet("Штрафы", data["violations"])
    sheet("Расхождения", data["discrepancies_detail"])
    sheet("Списания", data["writeoffs"])
    sheet("Зарплата", data["salaries"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = "strike_arena_statistics_all_time.xlsx" if days == 0 else f"strike_arena_statistics_{days}d.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})
